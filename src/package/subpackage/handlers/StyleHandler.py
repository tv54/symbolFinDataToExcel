import datetime
from openpyxl.styles import Font, Alignment
from openpyxl.chart.axis import DateAxis
from src.package.subpackage.othersrc.Constants import Constants

class StyleHandler():
    
    @classmethod
    def styleSheetConstantCols(cls, ws, mode, styles=None):
        match mode:
            case "profile":
                ws["B2"].font=Font(name='Calibri',size=18,bold=False)
                ws["G2"].font=Font(name='Calibri',size=18,bold=False)
                if(styles):
                    for i in range(Constants.COLUMN_NAMES.index("G")-Constants.COLUMN_NAMES.index("B")+1):
                        col=Constants.COLUMN_NAMES[i+1]
                        ws["{}5".format(col)].style=styles["summarySeparatorStyle"]
                        ws["{}15".format(col)].style=styles["summarySeparatorStyle"]
                        ws["{}22".format(col)].style=styles["summarySeparatorStyle"]
                        ws["{}29".format(col)].style=styles["summarySeparatorStyle"]
                        ws["{}37".format(col)].style=styles["summarySeparatorStyle"]
            case "cfs":
                ws["A2"].font=Font(size=16, bold=True)
                ws["A9"].font=Font(size=16, bold=True)
                ws["A13"].font=Font(size=16, bold=True)
                ws["A21"].font=Font(size=16, bold=True)
                ws["A23"].font=Font(size=16, bold=True)
                ws["A7"].alignment=Alignment(horizontal="center", vertical="center")
                ws["A11"].alignment=Alignment(horizontal="center", vertical="center")
                ws["A19"].alignment=Alignment(horizontal="center", vertical="center")
            case "bss":
                ws["A2"].font=Font(size=16, bold=True)
                ws["A12"].font=Font(size=16, bold=True)
                ws["A22"].font=Font(size=16, bold=True)
                ws["A3"].font=Font(size=14, bold=True)
                ws["A8"].font=Font(size=14, bold=True)
                ws["A13"].font=Font(size=14, bold=True)
                ws["A17"].font=Font(size=14, bold=True)
                ws["A7"].alignment=Alignment(horizontal="center", vertical="center")
                ws["A10"].alignment=Alignment(horizontal="center", vertical="center")
                ws["A16"].alignment=Alignment(horizontal="center", vertical="center")
                ws["A20"].alignment=Alignment(horizontal="center", vertical="center")
            case "is":
                ws["A2"].font=Font(size=14, bold=True)
                ws["A6"].font=Font(size=14, bold=True)
                ws["A9"].font=Font(size=14, bold=True)
                ws["A14"].font=Font(size=14, bold=True)
                ws["A17"].font=Font(size=16, bold=True)
                ws["A12"].alignment=Alignment(horizontal="center", vertical="center")

    @classmethod
    def styleSheet(cls, ws, mode, col="A"):
        match mode:
            case "profile":
                pass
            case "cfs":
                ws["{}2".format(col)].font=Font(size=16)
                ws["{}9".format(col)].font=Font(size=16)
                ws["{}13".format(col)].font=Font(size=16)
                ws["{}21".format(col)].font=Font(size=16)
                ws["{}23".format(col)].font=Font(size=16, bold=True)
                ws["{}7".format(col)].alignment=Alignment(horizontal="center", vertical="center")
                ws["{}11".format(col)].alignment=Alignment(horizontal="center", vertical="center")
                ws["{}19".format(col)].alignment=Alignment(horizontal="center", vertical="center")
            case "bss":
                ws["{}2".format(col)].font=Font(size=16)
                ws["{}12".format(col)].font=Font(size=16)
                ws["{}22".format(col)].font=Font(size=16)
                ws["{}3".format(col)].font=Font(size=14)
                ws["{}8".format(col)].font=Font(size=14)
                ws["{}13".format(col)].font=Font(size=14)
                ws["{}17".format(col)].font=Font(size=14)
                ws["{}7".format(col)].alignment=Alignment(horizontal="center", vertical="center")
                ws["{}10".format(col)].alignment=Alignment(horizontal="center", vertical="center")
                ws["{}16".format(col)].alignment=Alignment(horizontal="center", vertical="center")
                ws["{}20".format(col)].alignment=Alignment(horizontal="center", vertical="center")
            case "is":
                ws["{}2".format(col)].font=Font(size=14)
                ws["{}6".format(col)].font=Font(size=14)
                ws["{}9".format(col)].font=Font(size=14)
                ws["{}14".format(col)].font=Font(size=14)
                ws["{}17".format(col)].font=Font(size=16)
                ws["{}12".format(col)].alignment=Alignment(horizontal="center", vertical="center")

    @classmethod
    def cellFormatNumber(cls, ws, column, row, format="P", writeValue=""):
        match format.upper():
            case "D":
                ws["{}{}".format(column, row)].number_format="m/d/yyyy"
                if(writeValue):
                    try:
                        date=datetime.datetime.strptime(writeValue, "%Y-%m-%d")
                        return "=DATE({},{},{})".format(date.year, date.month, date.day)
                    except:
                        # input(writeValue)
                        pass
            case "P":
                ws["{}{}".format(column, row)].number_format="0.{}%".format("0" * Constants.ROUND_DECIMALS)
            case "A":
                ws["{}{}".format(column, row)].number_format=Constants.NUMBER_FORMAT_ACCOUNTING
        return ""

    @classmethod
    def defaultChartStyle(cls, chart, title, yAxisTitle, xAxisTitle, stacked, chartType="line", yearOnly=False, lineMarkers=False, series=0):
        chart.title=title
        chart.width=Constants.CHART_WIDTH
        chart.height=Constants.CHART_HEIGHT
        chart.marker=lineMarkers
        chart.legend.position=Constants.CHART_LEGEND_POSITION
        match chartType:
            case "line":
                chart.series[series].spPr.line.solidFill=Constants.DARK_BLUE_COLOR
            case "bar":
                chart.series[series].spPr.solidFill=Constants.DARK_BLUE_COLOR

        if(stacked):
            chart.grouping="stacked"
        # ---- x axis ----
        chart.x_axis=DateAxis(crossAx=100)
        chart.x_axis.title=xAxisTitle
        chart.x_axis.tickLblPos="low"
        if(yearOnly):
            chart.x_axis.numFmt="yyyy"
            chart.x_axis.baseTimeUnit="years"
        # ---- y axis ----
        chart.y_axis.title=yAxisTitle
        chart.y_axis.crossAx=500