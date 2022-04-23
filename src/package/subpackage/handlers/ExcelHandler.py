from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, PatternFill
from src.package.subpackage.handlers.ChartingHandler import ChartingHandler
from src.package.subpackage.othersrc.Constants import Constants
from src.package.subpackage.handlers.StyleHandler import StyleHandler
from src.package.subpackage.othersrc.genFunc import checkKeyInDict, handleError, handleInput, log, displayLine

class ExcelHandler(ChartingHandler):

    def __init__(self):
        self.wb=Workbook()
        self.styles={}
        self.__setDefaultSheets()
        self.__setDefaultStyles()
    
    def __setDefaultSheets(self):
        ws=self.wb.active 
        ws.sheet_properties.tabColor="FFFFFF"
        ws.sheet_view.showGridLines=Constants.SHOW_GRIDLINES

    def __setDefaultStyles(self):
        columnTitleStyle=NamedStyle(name="columnTitleStyle")
        columnTitleStyle.font=Font(name='Calibri',size=11,bold=True)
        columnTitleStyle.border=Border(bottom=Side(style='thick',color='000000'))
        columnTitleStyle.alignment=Alignment(horizontal="center", vertical="center")
        self.wb.add_named_style(columnTitleStyle)
        self.styles["columnTitleStyle"]=columnTitleStyle

        summarySeparatorStyle=NamedStyle(name="summarySeparatorStyle")
        summarySeparatorStyle.font=Font(name='Calibri',size=14,bold=False)
        summarySeparatorStyle.fill=PatternFill("solid", fgColor=Constants.LIGHT_GRAY_COLOR)
        self.wb.add_named_style(summarySeparatorStyle)
        self.styles["summarySeparatorStyle"]=summarySeparatorStyle

    def createSheet(self, sheetName, position=None, sheetColor="FFFFFF", showGridLines=Constants.SHOW_GRIDLINES):
        if(not position):
            ws=self.wb.create_sheet(sheetName, position)
        else:    
            ws=self.wb.create_sheet(sheetName)

        ws.sheet_properties.tabColor=sheetColor
        ws.sheet_view.showGridLines=showGridLines

        return ws
  
    def setColumnTitle(self, ws, column="A", columnTitle="", format=""):
        ws["{}1".format(column)].style=self.styles["columnTitleStyle"]

        if(format):
            retNum=StyleHandler.cellFormatNumber(ws, column, 1, format=format, writeValue=columnTitle)
            if(retNum):
                columnTitle=retNum

        ws["{}1".format(column)]=columnTitle

    def writeList(self, ws, list, listWriteValueKey="", column="A", startRow=2, columnTitle="", format="", isDic=True):
        columnTitle and self.setColumnTitle(ws, column, columnTitle)

        for i in range(len(list)):
            if(isDic and listWriteValueKey):
                writeValue=list[i][listWriteValueKey]
            else:
                writeValue=list[i]
                
            try:
                float(writeValue)
                writeValue=writeValue.__round__(Constants.ROUND_DECIMALS)
            except:
                pass

            retNum=StyleHandler.cellFormatNumber(ws, column, i+startRow, format=format, writeValue=writeValue)

            if(retNum):
                writeValue=retNum

            ws["{}{}".format(column, i+startRow)]=writeValue

    def saveWorkbook(self, wbName=""):
        if(not wbName):
            wbName="./" + Constants.OUTPUT_FOLDER + "/" + Constants.WORKBOOK_NAME
        elif(not wbName.__contains__(Constants.OUTPUT_FOLDER)):
            wbName="./" + Constants.OUTPUT_FOLDER + "/" + wbName
        if(checkKeyInDict(self.wb, "Sheet", False)):
            self.wb.remove("Sheet")
        log("Saving workbook {}".format(wbName))
        try:
            displayLine("Guardando Excel {}...".format(wbName), self.logFile)
            self.wb.save(wbName)
            displayLine("Guardando exitoso", self.logFile)
        except:
            handleError("Acceso al archivo {} denegado".format(wbName), self.logFile)
            userIn=handleInput("Nuevo nombre del Excel (enter para el mismo): ", self.logFile)
            if(userIn):
                if(not userIn.__contains__(".")):
                    userIn += ".xlsx"
                return self.saveWorkbook("./" + Constants.OUTPUT_FOLDER + "/" + userIn)
            else:
                return self.saveWorkbook(wbName)

        if(self.logFile):
            self.logFile.close()


