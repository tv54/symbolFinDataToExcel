import datetime
from openpyxl.styles import Alignment
from openpyxl.chart.reference import Reference
from openpyxl.worksheet.dimensions import SheetFormatProperties, ColumnDimension
from src.package.subpackage.othersrc.Constants import Constants
from src.package.subpackage.handlers.StyleHandler import StyleHandler
from src.package.subpackage.othersrc.genFunc import checkKeyInDict
from src.package.subpackage.handlers.ExcelHandler import ExcelHandler


class FinSheetsHandler(ExcelHandler):

    def __init__(self, n=0, benchmark="", bmkRes=None, priceToDisplay="close"):
        super().__init__()
        self.n=n
        self.benchmark=benchmark
        self.bmkRes=bmkRes
        self.priceToDisplay=priceToDisplay

    def __setDefaultValuesLists(self, ws, valuesDict, col="A", mode="cfs"):
        valuesList=[]
        StyleHandler.styleSheet(ws, mode, col)

        match mode:
            case "profile":
                valuesList=[
                    valuesDict["beta"],
                    valuesDict["volAvg"],
                    valuesDict["mktCap"]/1000,
                    valuesDict["industry"],
                    valuesDict["sector"],
                    valuesDict["website"],
                    valuesDict["lastDiv"],
                ]
            case "ratios":
                valuesList=[
                    valuesDict["priceEarningsRatioTTM"],
                    "=G2/G17",
                    valuesDict["priceToOperatingCashFlowsRatioTTM"],
                    valuesDict["enterpriseValueMultipleTTM"],
                    "",
                    "",
                    "",
                    valuesDict["currentRatioTTM"],
                    valuesDict["quickRatioTTM"],
                    valuesDict["cashRatioTTM"],
                    valuesDict["operatingCycleTTM"],
                    "",
                    "",
                    "",
                    valuesDict["operatingProfitMarginTTM"],
                    valuesDict["grossProfitMarginTTM"],
                    valuesDict["netProfitMarginTTM"],
                    valuesDict["returnOnEquityTTM"],
                    valuesDict["returnOnAssetsTTM"],
                    "",
                    "",
                    "",
                    valuesDict["debtEquityRatioTTM"],
                    valuesDict["interestCoverageTTM"],
                    valuesDict["cashFlowToDebtRatioTTM"]
                ]
            case "cfs":
                valuesList=[
                    valuesDict["netCashProvidedByOperatingActivities"]/Constants.DIV_OF_VALUES,
                    valuesDict["netIncome"]/Constants.DIV_OF_VALUES,
                    checkKeyInDict(valuesDict,"depreciationAndAmortization")*-1/Constants.DIV_OF_VALUES,
                    valuesDict["deferredIncomeTax"]/Constants.DIV_OF_VALUES,
                    valuesDict["changeInWorkingCapital"]/Constants.DIV_OF_VALUES,
                    "...",
                    "",
                    valuesDict["netCashUsedForInvestingActivites"]/Constants.DIV_OF_VALUES,
                    checkKeyInDict(valuesDict,"capitalExpenditure")/Constants.DIV_OF_VALUES,
                    "...",
                    "",
                    valuesDict["netCashUsedProvidedByFinancingActivities"]/Constants.DIV_OF_VALUES,
                    valuesDict["debtRepayment"]/Constants.DIV_OF_VALUES,
                    valuesDict["commonStockIssued"]/Constants.DIV_OF_VALUES,
                    valuesDict["commonStockRepurchased"]/Constants.DIV_OF_VALUES,
                    valuesDict["dividendsPaid"]/Constants.DIV_OF_VALUES,
                    valuesDict["otherFinancingActivites"]/Constants.DIV_OF_VALUES,
                    "...",
                    "",
                    "={0}2+{0}9+{0}13".format(col),
                    "",
                    "={0}2+{0}4".format(col)
                ]
            case "bss":
                valuesList=[
                    valuesDict["totalAssets"]/Constants.DIV_OF_VALUES,
                    valuesDict["totalCurrentAssets"]/Constants.DIV_OF_VALUES,
                    valuesDict["cashAndCashEquivalents"] + valuesDict["shortTermInvestments"]/Constants.DIV_OF_VALUES,
                    valuesDict["netReceivables"]/Constants.DIV_OF_VALUES,
                    valuesDict["inventory"]/Constants.DIV_OF_VALUES,
                    "...",
                    valuesDict["totalNonCurrentAssets"]/Constants.DIV_OF_VALUES,
                    valuesDict["propertyPlantEquipmentNet"]/Constants.DIV_OF_VALUES,
                    "...",
                    "",
                    valuesDict["totalLiabilities"]/Constants.DIV_OF_VALUES,
                    valuesDict["totalCurrentLiabilities"]/Constants.DIV_OF_VALUES,
                    valuesDict["shortTermDebt"]/Constants.DIV_OF_VALUES,
                    valuesDict["accountPayables"]/Constants.DIV_OF_VALUES,
                    "...",
                    valuesDict["totalNonCurrentLiabilities"]/Constants.DIV_OF_VALUES,
                    valuesDict["longTermDebt"]/Constants.DIV_OF_VALUES,
                    valuesDict["deferredTaxLiabilitiesNonCurrent"]/Constants.DIV_OF_VALUES,
                    "...",
                    "",
                    valuesDict["totalEquity"]/Constants.DIV_OF_VALUES,
                    valuesDict["commonStock"]/Constants.DIV_OF_VALUES,
                    valuesDict["retainedEarnings"]/Constants.DIV_OF_VALUES,
                    valuesDict["accumulatedOtherComprehensiveIncomeLoss"]/Constants.DIV_OF_VALUES,
                ]
            case "is":
                valuesList=[
                    valuesDict["grossProfit"]/Constants.DIV_OF_VALUES,
                    valuesDict["revenue"]/Constants.DIV_OF_VALUES,
                    valuesDict["costOfRevenue"]/Constants.DIV_OF_VALUES,
                    "",
                    valuesDict["operatingIncome"]/Constants.DIV_OF_VALUES,
                    valuesDict["operatingExpenses"]/Constants.DIV_OF_VALUES,
                    "",
                    valuesDict["incomeBeforeTax"]/Constants.DIV_OF_VALUES,
                    valuesDict["interestIncome"]/Constants.DIV_OF_VALUES,
                    valuesDict["interestExpense"]/Constants.DIV_OF_VALUES,
                    "...",
                    "",
                    valuesDict["netIncome"]/Constants.DIV_OF_VALUES,
                    valuesDict["incomeTaxExpense"]/Constants.DIV_OF_VALUES,
                    "",
                    valuesDict["ebitda"]/Constants.DIV_OF_VALUES,
                ]

        return valuesList

    def histPricesSheetBuilder(self, sym, res):
        ws=self.createSheet("Historicos")
        ws.sheet_format=SheetFormatProperties(defaultColWidth=Constants.HISTORICAL_SHEET_COL_WIDTH)

        chartCol="G"

        self.writeList(ws, res["historical"], "date", "A", columnTitle="", format="D")
        self.writeList(ws, res["historical"], self.priceToDisplay, "B", columnTitle="P {}".format(sym), format="A")
        if(self.bmkRes):
            self.writeList(ws, self.bmkRes["historical"], self.priceToDisplay, "G", columnTitle="P Benchmark", format="A")
            ws["K{}".format(self.n+2)]=Constants.BENCHMARK_PRICE_BASE_VALUE

        for i in range(self.n):
            # ---- Activo ----
            self.setColumnTitle(ws, "C", "r {}".format(sym))
            ws["C{}".format(i+2)].number_format="0.{}%".format("0" * Constants.ROUND_DECIMALS)
            ws["C{}".format(i+2)]="=IFERROR(B{}/B{}-1,0)".format(i+2,i+3)

            self.setColumnTitle(ws, "D", "Max")
            ws["D{}".format(i+2)].number_format=Constants.NUMBER_FORMAT_ACCOUNTING
            ws["D{}".format(i+2)]="=MAX($B{}:$B${})".format(i+2,self.n+1)

            self.setColumnTitle(ws, "E", "Caída")
            ws["E{}".format(i+2)].number_format="0.{}%".format("0" * Constants.ROUND_DECIMALS)
            ws["E{}".format(i+2)]="=$B{0}/$D{0}-1".format(i+2)           

            # ---- Benchmark ----
            if(self.bmkRes):
                chartCol="M"
                self.setColumnTitle(ws, "H", "r Benchmark")
                ws["H{}".format(i+2)].number_format="0.{}%".format("0" * Constants.ROUND_DECIMALS)
                ws["H{}".format(i+2)]="=IFERROR(G{}/G{}-1,0)".format(i+2,i+3)

                self.setColumnTitle(ws, "I", "Max")
                ws["I{}".format(i+2)].number_format=Constants.NUMBER_FORMAT_ACCOUNTING
                ws["I{}".format(i+2)]="=MAX($G{}:$G${})".format(i+2,self.n+1)

                self.setColumnTitle(ws, "J", "Caída")
                ws["J{}".format(i+2)].number_format="0.{}%".format("0" * Constants.ROUND_DECIMALS)
                ws["J{}".format(i+2)]="=$G{0}/$I{0}-1".format(i+2)           

                self.setColumnTitle(ws, "K", "Bmk (Base {})".format(Constants.BENCHMARK_PRICE_BASE_VALUE))
                ws["K{}".format(i+2)].number_format=Constants.NUMBER_FORMAT_ACCOUNTING
                ws["K{}".format(i+2)]="=$K{}*(1+H{})".format(i+3, i+2)           
                
        self.chart(ws, column1="A", column2="B", row1=1, row2=self.n+1, title="Crecimiento {}{}".format(sym, self.bmkRes and " vs benchmark {} (base {})".format(self.benchmark, Constants.BENCHMARK_PRICE_BASE_VALUE)), yAxisTitle="", chartCoords="{}3".format(chartCol), benchmarkCol="K", titlesFromData=True)
        if(checkKeyInDict(self.wb, "Datos {}".format(sym), False)):
            self.chart(ws, column1="A", column2="B", row1=1, row2=self.n+1, title="Crecimiento {}".format(sym), secondWs=self.wb["Datos {}".format(sym)], secondWsChartCoords="I2")
        self.chart(ws, column1="A", column2="C", row1=1, row2=self.n+1, title="r {}".format(sym), chartCoords="{}23".format(chartCol))
        self.chart(ws, column1="A", column2="E", row1=1, row2=self.n+1, title="{} drawdown{}".format(sym, self.bmkRes and " vs benchmark {} drawdown".format(self.benchmark)), chartCoords="{}43".format(chartCol), benchmarkCol="J")

    def cfsSheetBuilder(self, sym, res, limit):
        ws=self.createSheet("EFF {}".format(sym))
        ws.column_dimensions["A"]=ColumnDimension(ws, width=Constants.FINANCIAL_STATEMENTS_FIRST_COL_WIDTH)
        ws.sheet_format=SheetFormatProperties(defaultColWidth=Constants.FINANCIAL_STATEMENTS_OTHER_COL_WIDTH)

        StyleHandler.styleSheetConstantCols(ws, mode="cfs", styles=self.styles)
        self.writeList(ws, Constants.CFS_ROW_NAMES, isDic=False)

        for i in range(limit):
            val=res[i]
            col=Constants.COLUMN_NAMES[i+1]
            self.setColumnTitle(ws, col, val["date"], format="d")

            valuesList=self.__setDefaultValuesLists(ws, val, col, mode="cfs")
        
            self.writeList(ws, valuesList, column=col, isDic=False, format="A")

        self.chart(ws, column1="A", column2=Constants.COLUMN_NAMES[limit], row1=2, row2=1, title="Evolución del FEO", chartCoords="{}2".format(Constants.COLUMN_NAMES[limit+2]), lineMarkers=True, horizontalData=True)

    def bssSheetBuilder(self, sym, res, limit):
        ws=self.createSheet("BG {}".format(sym))
        ws.column_dimensions["A"]=ColumnDimension(ws, width=Constants.FINANCIAL_STATEMENTS_FIRST_COL_WIDTH)
        ws.sheet_format=SheetFormatProperties(defaultColWidth=Constants.FINANCIAL_STATEMENTS_OTHER_COL_WIDTH)
        
        StyleHandler.styleSheetConstantCols(ws, mode="bss", styles=self.styles)
        self.writeList(ws, Constants.BSS_ROW_NAMES, isDic=False)

        for i in range(limit):
            try:
                val=res[i]
                col=Constants.COLUMN_NAMES[i+1]
                self.setColumnTitle(ws, col, val["date"], format="d")

                valuesList=self.__setDefaultValuesLists(ws, val, col, mode="bss")
                
                self.writeList(ws, valuesList, column=col, isDic=False, format="A")
            except:
                pass

        self.chart(ws, column1="A", column2=Constants.COLUMN_NAMES[limit], row1=2, row2=1, title="Evolución del activo", chartCoords="H2", lineMarkers=True, horizontalData=True)

    def isSheetBuilder(self, sym, res, limit):
        ws=self.createSheet("EERR {}".format(sym))
        ws.column_dimensions["A"]=ColumnDimension(ws, width=Constants.FINANCIAL_STATEMENTS_FIRST_COL_WIDTH)
        ws.sheet_format=SheetFormatProperties(defaultColWidth=Constants.FINANCIAL_STATEMENTS_OTHER_COL_WIDTH)
        
        StyleHandler.styleSheetConstantCols(ws, mode="is", styles=self.styles)
        self.writeList(ws, Constants.IS_ROW_NAMES, isDic=False)

        for i in range(limit):
            val=res[i]
            col=Constants.COLUMN_NAMES[i+1]
            self.setColumnTitle(ws, col, val["date"], format="d")

            valuesList=self.__setDefaultValuesLists(ws, val, col, mode="is")
            
            self.writeList(ws, valuesList, column=col, isDic=False, format="A")

        self.chart(ws, column1="A", column2=Constants.COLUMN_NAMES[limit], row1=17, row2=1, title="Evolución del EBITDA", chartCoords="{}2".format(Constants.COLUMN_NAMES[limit+2]), lineMarkers=True, horizontalData=True)
        self.chart(ws, column1="A", column2=Constants.COLUMN_NAMES[limit], row1=3, row2=1, title="Evolución Ingresos totales y Resultado neto", horizontalData=True, chartCoords="{}22".format(Constants.COLUMN_NAMES[limit+2]), chartType="bar", extraElemReferences=[Reference(ws, min_col=1, max_col=limit+1, min_row=14)])
        if(checkKeyInDict(self.wb, "Datos {}".format(sym), False)):
            self.chart(ws, column1="A", column2=Constants.COLUMN_NAMES[limit], row1=3, row2=1, title="Evolución Ingresos totales y Resultado neto", horizontalData=True, chartCoords="", chartType="bar", extraElemReferences=[Reference(ws, min_col=1, max_col=limit+1, min_row=14)], secondWs=self.wb["Datos {}".format(sym)], secondWsChartCoords="I22")

    def summarySheetBuilder(self, sym, res1, res2):
        if(not checkKeyInDict(self.wb, "Sheet", False)):
            self.createSheet("Sheet")
        ws=self.wb["Sheet"]
        ws.title="Datos {}".format(sym.upper())
        ws.column_dimensions["G"]=ColumnDimension(ws, width=Constants.SUMMARY_NUMBER_COLUMN_WIDTH)
        ws.column_dimensions["A"]=ColumnDimension(ws, width=Constants.SUMMARY_FIRST_COLUMN_WIDTH)

        for i in range(41-7+1):
            StyleHandler.cellFormatNumber(ws, "G", i+7, "A")
            ws["G{}".format(i+7)].alignment=Alignment(horizontal="right")
        StyleHandler.cellFormatNumber(ws, "G", 3, "D")
        
        ws["B2"]="{} ({})".format(res1["companyName"], res1["currency"])
        ws["B3"]="{}:{}".format(res1["exchangeShortName"], res1["symbol"])
        ws["G2"]=res1["price"]

        date=datetime.datetime.now().date()
        ws["G3"]="=DATE({},{},{})".format(date.year, date.month, date.day)

        StyleHandler.styleSheetConstantCols(ws, mode="profile", styles=self.styles)
        self.writeList(ws, Constants.PROFILE_ROW_NAMES, isDic=False, startRow=5, column="B")

        valuesList=self.__setDefaultValuesLists(ws, res1, mode="profile")
        self.writeList(ws, valuesList, column="G", startRow=7, isDic=False)

        valuesList=self.__setDefaultValuesLists(ws, res2, mode="ratios")
        self.writeList(ws, valuesList, column="G", startRow=17, isDic=False)
